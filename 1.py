import openpyxl
from openpyxl.drawing.image import Image

# Open the Excel file
workbook = openpyxl.load_workbook('example.xlsx')

# Select the worksheet
worksheet = workbook['Sheet1']

# Loop through all cells in the worksheet
for row in worksheet.iter_rows():
    for cell in row:
        # Check if the cell has an image
        if cell.value and cell.value._type == 'image':
            # Extract the image from the cell
            img = Image(cell.value.blob)

            # Save the image to a file
            img.save('image.png')



import openpyxl
from PIL import Image

# Load the Excel file
workbook = openpyxl.load_workbook('example.xlsx')
worksheet = workbook['Sheet1']

# Iterate through all the cells in the worksheet
for row in worksheet.iter_rows():
    for cell in row:
        # Check if the cell contains an image
        if cell._value is not None and 'image' in cell._value:
            # Extract the image and save it to a specified directory
            img_path = 'images/' + cell._value.split('=')[1].strip("'")
            img_file = open(img_path, 'wb')
            img_file.write(cell.value)
            img_file.close()

# Display the successful completion of the task
print("All images have been extracted successfully.")

