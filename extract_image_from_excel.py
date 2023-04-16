import os
from openpyxl import load_workbook
from PIL import Image

input_file = "raw_data/Book1.xlsx"
output_folder = "images"

# Create the output folder if it doesn't exist
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# Load the input workbook
wb = load_workbook(input_file)

# Get the active sheet
ws = wb.active

# Define a translation table for replacing invalid characters
trans = str.maketrans({"/": "_", "\\": "_", ":": "_", "*": "_", "?": "_", "\"": "_", "<": "_", ">": "_", "|": "_"})

for img in ws._images:
    # Find the cell containing the image
    row = img.anchor._from.row + 1  # +1 because openpyxl row index is 0-based
    col = img.anchor._from.col + 1  # +1 because openpyxl col index is 0-based

    # Get the image filename from the value in column B (2)
    img_name = ws.cell(row=row, column=2).value

    # Replace invalid characters in the filename
    img_name = img_name.translate(trans)

    # Save the image in the output folder
    img_path = os.path.join(output_folder, f"{img_name}.png")
    pil_image = Image.open(img.ref)
    pil_image.save(img_path)





