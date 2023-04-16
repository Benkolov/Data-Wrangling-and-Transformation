import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image


class PipelineStep:
    def process(self, data):
        raise NotImplementedError("The process method must be implemented by subclasses")


class Pipeline:
    def __init__(self):
        self.steps = []

    def add_step(self, step):
        self.steps.append(step)

    def run(self, data):
        for step in self.steps:
            data = step.process(data)
        return data


class AddGroupColumn(PipelineStep):
    def process(self, data):
        data["Group"] = data["Description"].str.split(',').str[0]
        return data


class DataAggregation(PipelineStep):
    # def process(self, data):
    #     grouped_data = data.groupby("Group")
    #     for group_name, group_df in grouped_data:
    #         output_file = f"processed_data/{group_name}.xlsx"
    #         group_df.to_excel(output_file, index=False)
    #     return grouped_data
    def process(self, data):
        grouped_data = data.groupby("Group")
        input_file = "raw_data/Book1.xlsx"
        input_wb = load_workbook(input_file)
        input_ws = input_wb.active

        for group_name, group_df in grouped_data:
            output_file = f"processed_data/{group_name}.xlsx"
            group_df.to_excel(output_file, index=False)

            # Load the output file with openpyxl to add images
            output_wb = load_workbook(output_file)
            output_ws = output_wb.active

            # Iterate through rows and add images
            for index, row in group_df.iterrows():
                img_name = row['Item code']

                img_path = f"images/{img_name}.png"

                if os.path.exists(img_path):
                    img = Image(img_path)
                    input_row_num = data.index.get_loc(
                        index) + 2  # +2 because pandas index is 0-based and Excel is 1-based
                    output_row_num = output_ws.cell(row=output_ws.max_row, column=1).row + 1
                    output_ws.column_dimensions['A'].width = input_ws.column_dimensions['A'].width
                    output_ws.row_dimensions[output_row_num].height = input_ws.row_dimensions[input_row_num].height
                    output_ws.add_image(img, f"A{output_row_num}")

            # Save the output file with images
            output_wb.save(output_file)

        return grouped_data
