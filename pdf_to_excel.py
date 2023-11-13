import PyPDF2
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

pdf_path = "raw_data/SUMMER OFFERS 2023_web.pdf"

wb = Workbook()
ws = wb.active

pdf_file = open(pdf_path, 'rb')
pdf_reader = PyPDF2.PdfFileReader(pdf_file)

# Запис на текста в Excel файла
row_index = 1

for page_number, page in enumerate(pdf_reader.pages):
    text = page.extractText()

    lines = text.split('\n')
    for line in lines:
        # Извличане на желаните данни от всяка линия в PDF файла
        # Пример:
        for line in lines:
            if ':' in line:
                product_name, product_description = line.split(':', 1)
                product_name = product_name.strip()
                product_description = product_description.strip()
                # Запис на данните в Excel файла
                ws.cell(row=row_index, column=1, value=product_name)
                ws.cell(row=row_index, column=2, value=product_description)
            else:
                # Ако линията не съдържа ':', игнорирайте я или предприемете други действия по ваш избор
                continue



            row_index += 1

    # Запис на снимката в Excel файла
    image_path = f"image_{page_number}.png"
    with open(image_path, 'wb') as image_file:
        image_file.write(page["/Resources"]["/XObject"].getObject()["/Im0"].getData())

    img = PILImage.open(image_path)
    img.thumbnail((200, 200))  # Променете размера на снимката според вашите нужди
    excel_image = Image(img)
    ws.column_dimensions['C'].width = 40  # Настройване на ширината на колоната за изображения
    ws.row_dimensions[row_index - 1].height = 100  # Настройване на височината на реда за изображения
    ws.add_image(excel_image, f"C{row_index - 1}")


excel_file_path = "product_catalog.xlsx"
wb.save(excel_file_path)
